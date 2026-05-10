"""
读取 results/ 下的CSV, 生成漂亮的 xlsx 报告
"""
import os
import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

RESULTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "results")


def _hdr(cell):
    cell.font = Font(name="Arial", bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _border(cell):
    thin = Side(border_style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def build_summary_sheet(ws, summary_rows):
    headers = ["排名", "策略变体", "交易次数", "胜次", "胜率", "Total R", "Avg R/笔",
               "最大回撤(R)", "Body阈值", "盈亏比 R", "SL缓冲%", "时间止损", "保本上移"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        _hdr(c)
    # 按 Total R 降序
    sorted_rows = sorted(summary_rows, key=lambda x: float(x["total_r"]), reverse=True)
    for i, row in enumerate(sorted_rows, 2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=row["variant"])
        ws.cell(row=i, column=3, value=int(row["n_trades"]))
        ws.cell(row=i, column=4, value=int(row["n_wins"]))
        wr_cell = ws.cell(row=i, column=5, value=float(row["win_rate"]))
        wr_cell.number_format = "0.0%"
        tr_cell = ws.cell(row=i, column=6, value=float(row["total_r"]))
        tr_cell.number_format = "+0.00;-0.00;-"
        ar_cell = ws.cell(row=i, column=7, value=float(row["avg_r"]))
        ar_cell.number_format = "+0.000;-0.000;-"
        dd_cell = ws.cell(row=i, column=8, value=float(row["max_dd_r"]))
        dd_cell.number_format = "0.00"
        ws.cell(row=i, column=9, value=float(row["body_ratio"]))
        ws.cell(row=i, column=10, value=f"{row['r_multiple']}R")
        ws.cell(row=i, column=11, value=f"{float(row['sl_buffer_pct'])*100:.1f}%")
        ws.cell(row=i, column=12, value=int(row["time_stop_bars"]) if int(row["time_stop_bars"]) else "-")
        ws.cell(row=i, column=13, value=f"{row['breakeven_at_r']}R" if float(row["breakeven_at_r"]) else "-")
        for col in range(1, 14):
            cc = ws.cell(row=i, column=col)
            cc.font = Font(name="Arial")
            _border(cc)
            if col in (1, 3, 4, 12):
                cc.alignment = Alignment(horizontal="center")
    # 给Total R和WinRate加色阶
    if len(sorted_rows) > 1:
        last = len(sorted_rows) + 1
        ws.conditional_formatting.add(
            f"E2:E{last}",
            ColorScaleRule(start_type="min", start_color="F8696B",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="max", end_color="63BE7B"),
        )
        ws.conditional_formatting.add(
            f"F2:F{last}",
            ColorScaleRule(start_type="min", start_color="F8696B",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="max", end_color="63BE7B"),
        )
        # 回撤 反向色阶
        ws.conditional_formatting.add(
            f"H2:H{last}",
            ColorScaleRule(start_type="min", start_color="63BE7B",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="max", end_color="F8696B"),
        )
    # 列宽
    widths = [6, 22, 9, 7, 8, 9, 10, 12, 10, 9, 14, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def build_matrix_sheet(ws, matrix_path):
    if not os.path.exists(matrix_path):
        return
    with open(matrix_path) as f:
        reader = list(csv.reader(f))
    if not reader:
        return
    for ri, row in enumerate(reader, 1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Arial", bold=(ri == 1 or ci == 1))
            _border(cell)
            if ri == 1:
                _hdr(cell)
            elif ci > 1:
                cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 24
    for c in range(2, len(reader[0]) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13
    ws.freeze_panes = "B2"


def build_trades_sheet(ws, trades_path):
    if not os.path.exists(trades_path):
        return
    with open(trades_path) as f:
        reader = list(csv.reader(f))
    if not reader:
        return
    for ri, row in enumerate(reader, 1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Arial")
            if ri == 1:
                _hdr(cell)
            else:
                _border(cell)
                # 净R着色
                if ri > 1 and ci == 12:
                    try:
                        v = float(val)
                        if v > 0:
                            cell.fill = PatternFill("solid", start_color="E8F5E9")
                            cell.font = Font(name="Arial", color="2E7D32")
                        elif v < 0:
                            cell.fill = PatternFill("solid", start_color="FFEBEE")
                            cell.font = Font(name="Arial", color="C62828")
                    except Exception:
                        pass
                if ci == 13:  # win列
                    if val == "True":
                        cell.font = Font(name="Arial", color="2E7D32", bold=True)
                    elif val == "False":
                        cell.font = Font(name="Arial", color="C62828")
    ws.freeze_panes = "A2"
    for c in range(1, 14):
        ws.column_dimensions[get_column_letter(c)].width = 12


def build_overview(ws, summary_rows):
    ws.cell(row=1, column=1, value="两线反转策略 — 赛马回测报告").font = Font(name="Arial", bold=True, size=16)
    ws.merge_cells("A1:E1")
    ws.cell(row=2, column=1, value="基于 GiantCutie 的 TradingView 脚本 v2 改写, R-Multiple 回测").font = Font(name="Arial", italic=True, color="555555")
    ws.merge_cells("A2:E2")

    sorted_by_total = sorted(summary_rows, key=lambda x: float(x["total_r"]), reverse=True)
    sorted_by_winrate = sorted(summary_rows, key=lambda x: float(x["win_rate"]), reverse=True)
    sorted_by_avg = sorted(summary_rows, key=lambda x: float(x["avg_r"]), reverse=True)

    blocks = [
        ("总R榜首 (累计盈利能力)", sorted_by_total[:3]),
        ("胜率榜首 (高确定性)", sorted_by_winrate[:3]),
        ("Avg R/笔 榜首 (单笔效率)", sorted_by_avg[:3]),
    ]
    row = 4
    for title, top3 in blocks:
        ws.cell(row=row, column=1, value=title).font = Font(name="Arial", bold=True, size=12, color="1F4E78")
        row += 1
        ws.cell(row=row, column=1, value="排名").font = Font(name="Arial", bold=True)
        ws.cell(row=row, column=2, value="变体").font = Font(name="Arial", bold=True)
        ws.cell(row=row, column=3, value="交易数").font = Font(name="Arial", bold=True)
        ws.cell(row=row, column=4, value="胜率").font = Font(name="Arial", bold=True)
        ws.cell(row=row, column=5, value="Total R").font = Font(name="Arial", bold=True)
        row += 1
        for rank, r in enumerate(top3, 1):
            ws.cell(row=row, column=1, value=rank)
            ws.cell(row=row, column=2, value=r["variant"])
            ws.cell(row=row, column=3, value=int(r["n_trades"]))
            c4 = ws.cell(row=row, column=4, value=float(r["win_rate"])); c4.number_format = "0.0%"
            c5 = ws.cell(row=row, column=5, value=float(r["total_r"])); c5.number_format = "+0.00;-0.00;-"
            row += 1
        row += 1

    ws.cell(row=row + 1, column=1, value="说明:").font = Font(name="Arial", bold=True)
    notes = [
        "1) 止损: 空 SL = max(B高, C高) × (1+buffer); 多 SL = min(B低, C低) × (1-buffer)。默认 buffer=2%。",
        "2) R = |入场价 - SL|, 止盈 = 入场 ± N×R (N 由变体决定: 1.5/2/2.5/3)。",
        "3) 入场: 信号K线的下一根开盘价。同根K线SL+TP都触及时, 保守按SL先算。",
        "4) 手续费按 0.05% 单边 (合约 taker 参考), 已在 net_r 中扣除。",
        "5) Total R 是该变体跨所有交易的累计 R 收益, 越高越好。Max DD 是 R 单位的最大回撤。",
        "6) 注意小样本陷阱: 信号数 <30 的变体胜率波动大, 不要轻信。",
    ]
    for i, n in enumerate(notes):
        ws.cell(row=row + 2 + i, column=1, value=n).font = Font(name="Arial", color="333333")
        ws.merge_cells(start_row=row + 2 + i, start_column=1, end_row=row + 2 + i, end_column=8)

    for col, w in zip("ABCDE", [6, 26, 10, 10, 10]):
        ws.column_dimensions[col].width = w


def main():
    summary_csv = os.path.join(RESULTS_DIR, "summary.csv")
    if not os.path.exists(summary_csv):
        print(f"ERROR: 找不到 {summary_csv} - 请先运行 run_backtest.py")
        return
    with open(summary_csv) as f:
        summary_rows = list(csv.DictReader(f))

    wb = Workbook()
    ws_overview = wb.active
    ws_overview.title = "总览"
    build_overview(ws_overview, summary_rows)

    ws_summary = wb.create_sheet("策略变体汇总")
    build_summary_sheet(ws_summary, summary_rows)

    ws_matrix = wb.create_sheet("胜率矩阵")
    build_matrix_sheet(ws_matrix, os.path.join(RESULTS_DIR, "winrate_matrix.csv"))

    ws_trades = wb.create_sheet("交易明细")
    build_trades_sheet(ws_trades, os.path.join(RESULTS_DIR, "all_trades.csv"))

    out = os.path.join(RESULTS_DIR, "回测报告.xlsx")
    wb.save(out)
    print(f"报告生成: {out}")


if __name__ == "__main__":
    main()
